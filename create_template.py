"""
Genera horarios_data.xlsx — plantilla editable para cargar turnos.
Corré este script para crear/resetear la plantilla.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db import get_db, init_db

HEADER_FILL  = PatternFill("solid", fgColor="1A1D27")
HEADER_FONT  = Font(bold=True, color="E8EAF0", size=10)
LABEL_FILLS  = {
    "AIRE":       "0E3A5C",
    "EDICION":    "2D1B5C",
    "ZOCALOS":    "0C3B2A",
    "PLACAS":     "5C2A0C",
    "TEXTOS":     "5C1B3A",
    "CONTENIDOS": "5C4A0C",
}
THIN = Side(style="thin", color="2E3350")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr(ws, row, col, val, fill_hex=None, bold=True, color="E8EAF0", align="center", size=10):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, color=color, size=size)
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell

def create_template():
    wb = Workbook()

    # ── TURNOS ────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "TURNOS"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"

    headers = ["SEMANA", "AÑO", "FECHA", "FUNCION", "CANAL",
               "SHOW_INICIO", "SHOW_FIN", "EMPLEADO", "INGRESO", "EGRESO", "TIPO"]
    col_widths = [9, 6, 12, 12, 14, 12, 12, 22, 8, 8, 9]
    notes = [
        "Nro semana ISO (1-52)",
        "2026",
        "DD/MM/AAAA",
        "AIRE / EDICION / ZOCALOS / PLACAS / TEXTOS / CONTENIDOS",
        "ESPN / ESPN 2/ESPN3 / CHI / COL / CAM / REGIONES",
        "HH:MM",
        "HH:MM",
        "Apellido Nombre",
        "HH:MM",
        "HH:MM",
        "trabajo / libre",
    ]

    # Row 1: notes (gray)
    for c, note in enumerate(notes, 1):
        cell = ws.cell(row=1, column=c, value=note)
        cell.font = Font(color="8891AA", size=8, italic=True)
        cell.alignment = Alignment(horizontal="center")

    # Row 2: headers
    for c, h in enumerate(headers, 1):
        hdr(ws, 2, c, h, fill_hex="0F1117")

    # Example rows
    examples = [
        [1, 2026, "29/12/2025", "AIRE",    "ESPN",         "12:00", "14:00", "ROBERTO PACHECO",   "08:00", "16:00", "trabajo"],
        [1, 2026, "29/12/2025", "AIRE",    "ESPN",         "14:00", "16:30", "MATIAS DI NATALE",  "10:00", "18:00", "trabajo"],
        [1, 2026, "29/12/2025", "EDICION", "ESPN 2/ESPN3", "10:00", "12:00", "LUCAS PARNES",      "08:00", "16:00", "trabajo"],
        [1, 2026, "29/12/2025", "",        "",             "",      "",      "GIMENA CARBALLO",   "",      "",      "libre"],
    ]
    for r, row in enumerate(examples, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            fn = row[3] if len(row) > 3 else ""
            if fn and fn in LABEL_FILLS:
                cell.fill = PatternFill("solid", fgColor=LABEL_FILLS[fn])
                cell.font = Font(color="E8EAF0", size=9)
            else:
                cell.font = Font(color="E8EAF0", size=9)
                cell.fill = PatternFill("solid", fgColor="1A1D27")

    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    # ── EMPLEADOS ─────────────────────────────────────────────────────────────
    we = wb.create_sheet("EMPLEADOS")
    we.sheet_view.showGridLines = False
    emp_headers = ["NOMBRE", "FUNCION", "EMPRESA"]
    emp_widths  = [25, 15, 15]
    for c, h in enumerate(emp_headers, 1):
        hdr(we, 1, c, h, fill_hex="0F1117")

    # Pull from DB
    init_db()
    db = get_db()
    emps = db.execute("SELECT nombre, funcion, empresa FROM empleados ORDER BY nombre").fetchall()
    db.close()

    for r, emp in enumerate(emps, 2):
        for c, val in enumerate([emp['nombre'], emp['funcion'], emp['empresa']], 1):
            cell = we.cell(row=r, column=c, value=val)
            cell.font = Font(color="E8EAF0", size=9)
            cell.fill = PatternFill("solid", fgColor="1A1D27")
            cell.border = BORDER
    for c, w in enumerate(emp_widths, 1):
        we.column_dimensions[get_column_letter(c)].width = w

    # ── FERIADOS ──────────────────────────────────────────────────────────────
    wf = wb.create_sheet("FERIADOS")
    wf.sheet_view.showGridLines = False
    fer_headers = ["FECHA", "DESCRIPCION"]
    fer_widths  = [14, 28]
    for c, h in enumerate(fer_headers, 1):
        hdr(wf, 1, c, h, fill_hex="0F1117")
    feriados_ejemplo = [
        ["01/01/2026", "Año Nuevo"],
        ["16/02/2026", "Carnaval"],
        ["17/02/2026", "Carnaval"],
        ["20/03/2026", "Día de la Memoria"],
        ["24/03/2026", "Día Nacional de la Memoria"],
        ["02/04/2026", "Malvinas"],
        ["03/04/2026", "Viernes Santo"],
        ["01/05/2026", "Día del Trabajo"],
        ["25/05/2026", "Revolución de Mayo"],
        ["20/06/2026", "Día de la Bandera"],
        ["09/07/2026", "Independencia"],
        ["10/07/2026", "Puente turístico"],
        ["17/08/2026", "San Martín"],
        ["12/11/2026", "Día de la Diversidad Cultural"],
        ["02/11/2026", "Feriado puente"],
        ["07/12/2026", "Inmaculada Concepción"],
        ["08/12/2026", "Inmaculada Concepción"],
        ["25/12/2026", "Navidad"],
    ]
    for r, row in enumerate(feriados_ejemplo, 2):
        for c, val in enumerate(row, 1):
            cell = wf.cell(row=r, column=c, value=val)
            cell.font = Font(color="E8EAF0", size=9)
            cell.fill = PatternFill("solid", fgColor="1A1D27")
            cell.border = BORDER
    for c, w in enumerate(fer_widths, 1):
        wf.column_dimensions[get_column_letter(c)].width = w

    path = "horarios_data.xlsx"
    wb.save(path)
    print(f"Plantilla creada: {path}")
    print(f"  - {len(emps)} empleados exportados")
    print(f"  - {len(feriados_ejemplo)} feriados de ejemplo")

if __name__ == "__main__":
    create_template()
